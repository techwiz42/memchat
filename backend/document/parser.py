"""Extract plain text from uploaded document files."""

import csv
import io

import pdfplumber
from docx import Document as DocxDocument
from openpyxl import load_workbook

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".tiff", ".tif"}

ALLOWED_EXTENSIONS = {".txt", ".md", ".pdf", ".docx", ".xlsx", ".csv"} | IMAGE_EXTENSIONS


async def extract_text(filename: str, content: bytes) -> str:
    """Extract plain text from a file based on its extension.

    Args:
        filename: Original filename (used to determine type).
        content: Raw file bytes.

    Returns:
        Extracted text as a string.

    Raises:
        ValueError: If the file type is not supported.
    """
    ext = _get_extension(filename)

    if ext in (".txt", ".md"):
        return content.decode("utf-8", errors="replace")

    if ext == ".pdf":
        return _extract_pdf(content)

    if ext == ".docx":
        return _extract_docx(content)

    if ext == ".xlsx":
        return _extract_xlsx(content)

    if ext == ".csv":
        return _extract_csv(content)

    raise ValueError(
        f"Unsupported file type: {ext}. "
        f"Allowed types: {', '.join(sorted(ALLOWED_EXTENSIONS))}"
    )


def _get_extension(filename: str) -> str:
    """Return the lowercase file extension including the dot."""
    dot = filename.rfind(".")
    if dot == -1:
        return ""
    return filename[dot:].lower()


def _extract_pdf(content: bytes) -> str:
    """Extract text from a PDF using pdfplumber."""
    pages = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
    return "\n\n".join(pages)


def _extract_docx(content: bytes) -> str:
    """Extract text from a DOCX file using python-docx."""
    doc = DocxDocument(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def _extract_xlsx(content: bytes) -> str:
    """Extract text from an XLSX file using openpyxl."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts)


def _extract_csv(content: bytes) -> str:
    """Extract text from a CSV file."""
    text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = []
    for row in reader:
        if any(cell.strip() for cell in row):
            rows.append("\t".join(row))
    return "\n".join(rows)
